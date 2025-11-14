"""
Módulo para geração de relatórios Excel de análise de risco de endereço.
"""

from datetime import datetime
from typing import Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO


def formatar_cnpj(cnpj: str) -> str:
    """Formata CNPJ para XX.XXX.XXX/XXXX-XX."""
    cnpj_clean = "".join(filter(str.isdigit, cnpj))
    if len(cnpj_clean) == 14:
        return f"{cnpj_clean[:2]}.{cnpj_clean[2:5]}.{cnpj_clean[5:8]}/{cnpj_clean[8:12]}-{cnpj_clean[12:]}"
    return cnpj_clean


def gerar_relatorio_excel(
    cnpj: str,
    dados_empresa: Dict[str, Any],
    dados_endereco: Optional[Dict[str, Any]],
    analise_risco: Dict[str, Any],
    cnaes: list,
    caminho_saida: Optional[str] = None
) -> bytes:
    """
    Gera relatório Excel completo de análise de risco.
    
    Args:
        cnpj: CNPJ da empresa
        dados_empresa: Dados da empresa (razão social, nome fantasia, etc.)
        dados_endereco: Dados de geocoding e endereço
        analise_risco: Resultado completo da análise de risco
        cnaes: Lista de CNAEs da empresa
        caminho_saida: Caminho para salvar o arquivo (opcional, retorna bytes se None)
    
    Returns:
        Bytes do arquivo Excel ou None se caminho_saida fornecido
    """
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análise de Risco"
    
    # Estilos
    estilo_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    estilo_cabecalho = Font(name="Arial", size=11, bold=True)
    estilo_normal = Font(name="Arial", size=10)
    estilo_risco_alto = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    estilo_risco_medio = Font(name="Arial", size=11, bold=True, color="000000")
    estilo_risco_baixo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    fill_titulo = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_cabecalho = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_risco_alto = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_risco_medio = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    fill_risco_baixo = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    fill_cinza = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_esquerda = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    linha_atual = 1
    
    # TÍTULO
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_titulo = ws[f"A{linha_atual}"]
    celula_titulo.value = "RELATÓRIO DE ANÁLISE DE RISCO DE ENDEREÇO"
    celula_titulo.font = estilo_titulo
    celula_titulo.fill = fill_titulo
    celula_titulo.alignment = alinhamento_centro
    linha_atual += 1
    
    # Data de geração
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_data = ws[f"A{linha_atual}"]
    celula_data.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    celula_data.font = Font(name="Arial", size=9, italic=True)
    celula_data.alignment = alinhamento_centro
    linha_atual += 2
    
    # SEÇÃO 1: DADOS DA EMPRESA
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec1 = ws[f"A{linha_atual}"]
    celula_sec1.value = "1. DADOS DA EMPRESA"
    celula_sec1.font = estilo_cabecalho
    celula_sec1.fill = fill_cabecalho
    celula_sec1.alignment = alinhamento_esquerda
    linha_atual += 1
    
    # Dados da empresa
    email_cadastrado_display = dados_empresa.get("email_cadastrado") or "Não informado"
    email_cnpja_display = dados_empresa.get("email_cnpja") or "Não encontrado"
    
    dados_empresa_lista = [
        ("CNPJ", formatar_cnpj(cnpj)),
        ("Razão Social", dados_empresa.get("razao_social", "N/A")),
        ("Nome Fantasia", dados_empresa.get("nome_fantasia", "N/A")),
        ("Data de Abertura", dados_empresa.get("data_abertura", "N/A")),
        ("Email Cadastrado", email_cadastrado_display),
        ("Email CNPJA", email_cnpja_display),
    ]
    
    for rotulo, valor in dados_empresa_lista:
        ws[f"A{linha_atual}"] = rotulo
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        ws[f"A{linha_atual}"].alignment = alinhamento_esquerda
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        ws[f"B{linha_atual}"] = valor
        ws[f"B{linha_atual}"].font = estilo_normal
        ws[f"B{linha_atual}"].border = border
        ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
        linha_atual += 1
    
    linha_atual += 1
    
    # SCORE DE RISCO EM DESTAQUE (logo após dados da empresa)
    risco_final = analise_risco.get("risco_final", "INDEFINIDO")
    score_risco = analise_risco.get("score_risco", 0)
    
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_score_titulo = ws[f"A{linha_atual}"]
    celula_score_titulo.value = "SCORE DE RISCO"
    celula_score_titulo.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    celula_score_titulo.fill = fill_titulo
    celula_score_titulo.alignment = alinhamento_centro
    linha_atual += 1
    
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_score_valor = ws[f"A{linha_atual}"]
    celula_score_valor.value = f"{score_risco}/100"
    
    # Aplicar cores e estilos baseados no score
    if risco_final == "ALTO" or score_risco >= 60:
        celula_score_valor.font = Font(name="Arial", size=24, bold=True, color="FFFFFF")
        celula_score_valor.fill = fill_risco_alto
    elif risco_final == "MEDIO" or score_risco >= 30:
        celula_score_valor.font = Font(name="Arial", size=24, bold=True, color="000000")
        celula_score_valor.fill = fill_risco_medio
    elif risco_final == "BAIXO":
        celula_score_valor.font = Font(name="Arial", size=24, bold=True, color="FFFFFF")
        celula_score_valor.fill = fill_risco_baixo
    else:
        celula_score_valor.font = Font(name="Arial", size=24, bold=True, color="FFFFFF")
        celula_score_valor.fill = fill_cinza
    
    celula_score_valor.alignment = alinhamento_centro
    celula_score_valor.border = border
    linha_atual += 1
    
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_risco_status = ws[f"A{linha_atual}"]
    celula_risco_status.value = f"Risco: {risco_final}"
    celula_risco_status.font = Font(name="Arial", size=12, bold=True)
    celula_risco_status.alignment = alinhamento_centro
    celula_risco_status.border = border
    
    if risco_final == "ALTO":
        celula_risco_status.fill = fill_risco_alto
        celula_risco_status.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    elif risco_final == "MEDIO":
        celula_risco_status.fill = fill_risco_medio
        celula_risco_status.font = Font(name="Arial", size=12, bold=True, color="000000")
    elif risco_final == "BAIXO":
        celula_risco_status.fill = fill_risco_baixo
        celula_risco_status.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    else:
        celula_risco_status.fill = fill_cinza
    
    linha_atual += 2
    
    # SEÇÃO 2: ENDEREÇO
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec2 = ws[f"A{linha_atual}"]
    celula_sec2.value = "2. ENDEREÇO"
    celula_sec2.font = estilo_cabecalho
    celula_sec2.fill = fill_cabecalho
    celula_sec2.alignment = alinhamento_esquerda
    linha_atual += 1
    
    if dados_endereco:
        endereco_formatado = dados_endereco.get("formatted_address") or dados_endereco.get("endereco_completo", "N/A")
        lat = dados_endereco.get("lat")
        lng = dados_endereco.get("lng")
        
        dados_endereco_lista = [
            ("Endereço Completo", endereco_formatado),
            ("Coordenadas", f"{lat}, {lng}" if lat and lng else "N/A"),
            ("Place ID", dados_endereco.get("place_id", "N/A")),
        ]
        
        for rotulo, valor in dados_endereco_lista:
            ws[f"A{linha_atual}"] = rotulo
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            ws[f"A{linha_atual}"].alignment = alinhamento_esquerda
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = valor
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
        
        # Adicionar imagem do Google Maps (Street View ou Places)
        imagem_bytes = None
        tipo_imagem = None
        
        # Prioridade 1: Street View
        if dados_endereco.get("street_view_image_bytes"):
            imagem_bytes = dados_endereco.get("street_view_image_bytes")
            tipo_imagem = "Street View"
        # Prioridade 2: Fotos do Places (primeira foto disponível)
        elif dados_endereco.get("place_photos"):
            place_photos = dados_endereco.get("place_photos", [])
            if place_photos and len(place_photos) > 0:
                primeira_foto = place_photos[0]
                if isinstance(primeira_foto, dict) and primeira_foto.get("image_bytes"):
                    imagem_bytes = primeira_foto.get("image_bytes")
                    tipo_imagem = "Google Places"
        
        # Inserir imagem se disponível
        if imagem_bytes:
            linha_atual += 1
            
            # Título da imagem
            ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
            celula_img_titulo = ws[f"A{linha_atual}"]
            celula_img_titulo.value = f"Imagem do Endereço ({tipo_imagem})"
            celula_img_titulo.font = estilo_cabecalho
            celula_img_titulo.fill = fill_cinza
            celula_img_titulo.border = border
            celula_img_titulo.alignment = alinhamento_esquerda
            linha_atual += 1
            
            # Inserir imagem
            try:
                # Converter bytes para objeto Image do openpyxl
                img = Image(BytesIO(imagem_bytes))
                
                # Redimensionar imagem para caber no Excel
                # openpyxl usa pixels, mas Excel trabalha melhor com tamanhos menores
                # Largura máxima ~600 pixels, altura máxima ~400 pixels
                max_width = 600
                max_height = 400
                
                if img.width > max_width or img.height > max_height:
                    ratio = min(max_width / img.width, max_height / img.height)
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                
                # Posicionar imagem (centro das colunas B, C, D)
                img.anchor = f"B{linha_atual}"
                
                # Adicionar imagem à planilha
                ws.add_image(img)
                
                # Ajustar altura da linha para acomodar a imagem
                # Excel usa pontos (points) para altura de linha: 1 ponto ≈ 1.33 pixels
                # Converter pixels para pontos e adicionar margem
                altura_linha_pontos = max(30, int(img.height / 1.33) + 10)
                ws.row_dimensions[linha_atual].height = altura_linha_pontos
                
                # Calcular quantas linhas adicionais a imagem pode ocupar
                # (a imagem pode se estender por múltiplas linhas)
                linhas_adicionais = max(1, int((img.height / 1.33) / 15))  # ~15 pontos por linha padrão
                linha_atual += linhas_adicionais
                
            except Exception as e:
                # Se houver erro ao inserir imagem, apenas registrar
                ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
                ws[f"A{linha_atual}"] = f"Erro ao carregar imagem: {str(e)}"
                ws[f"A{linha_atual}"].font = Font(name="Arial", size=9, italic=True, color="FF0000")
                linha_atual += 1
    else:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        ws[f"A{linha_atual}"] = "Endereço não processado"
        ws[f"A{linha_atual}"].font = estilo_normal
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 3: CNAEs
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec3 = ws[f"A{linha_atual}"]
    celula_sec3.value = "3. ATIVIDADES CNAE"
    celula_sec3.font = estilo_cabecalho
    celula_sec3.fill = fill_cabecalho
    celula_sec3.alignment = alinhamento_esquerda
    linha_atual += 1
    
    # Cabeçalho da tabela CNAE
    ws[f"A{linha_atual}"] = "Tipo"
    ws[f"B{linha_atual}"] = "Código CNAE"
    ws.merge_cells(f"C{linha_atual}:D{linha_atual}")
    ws[f"C{linha_atual}"] = "Descrição"
    
    for col in ["A", "B", "C", "D"]:
        celula = ws[f"{col}{linha_atual}"]
        celula.font = estilo_cabecalho
        celula.fill = fill_cabecalho
        celula.border = border
        celula.alignment = alinhamento_centro
    
    linha_atual += 1
    
    # CNAE Principal
    if cnaes and len(cnaes) > 0:
        cnae_principal = cnaes[0]
        ws[f"A{linha_atual}"] = "Principal"
        ws[f"B{linha_atual}"] = cnae_principal.get("codigo", "N/A")
        ws.merge_cells(f"C{linha_atual}:D{linha_atual}")
        ws[f"C{linha_atual}"] = cnae_principal.get("descricao", "N/A")
        
        for col in ["A", "B", "C", "D"]:
            celula = ws[f"{col}{linha_atual}"]
            celula.font = estilo_normal
            celula.border = border
            celula.alignment = alinhamento_esquerda if col != "A" else alinhamento_centro
        
        linha_atual += 1
        
        # CNAEs Secundários
        if len(cnaes) > 1:
            for cnae_sec in cnaes[1:]:
                ws[f"A{linha_atual}"] = "Secundária"
                ws[f"B{linha_atual}"] = cnae_sec.get("codigo", "N/A")
                ws.merge_cells(f"C{linha_atual}:D{linha_atual}")
                ws[f"C{linha_atual}"] = cnae_sec.get("descricao", "N/A")
                
                for col in ["A", "B", "C", "D"]:
                    celula = ws[f"{col}{linha_atual}"]
                    celula.font = estilo_normal
                    celula.border = border
                    celula.alignment = alinhamento_esquerda if col != "A" else alinhamento_centro
                
                linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 3.5: ANÁLISE SEMÂNTICA DE CNAE (IA)
    avaliacao_cnae = dados_empresa.get("avaliacao_cnae")
    if avaliacao_cnae:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_sec35 = ws[f"A{linha_atual}"]
        celula_sec35.value = "3.5. ANÁLISE SEMÂNTICA DE CNAE (IA)"
        celula_sec35.font = estilo_cabecalho
        celula_sec35.fill = fill_cabecalho
        celula_sec35.alignment = alinhamento_esquerda
        linha_atual += 1
        
        # Compatível
        compativel = avaliacao_cnae.get("compativel")
        ws[f"A{linha_atual}"] = "Compatível:"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        celula_compativel = ws[f"B{linha_atual}"]
        if compativel is True:
            celula_compativel.value = "✅ Sim"
            celula_compativel.font = Font(name="Arial", size=10, color="006100", bold=True)
        elif compativel is False:
            celula_compativel.value = "❌ Não"
            celula_compativel.font = Font(name="Arial", size=10, color="C00000", bold=True)
        else:
            celula_compativel.value = "Indefinido"
            celula_compativel.font = estilo_normal
        celula_compativel.border = border
        linha_atual += 1
        
        # Score
        score = avaliacao_cnae.get("score")
        if score is not None:
            ws[f"A{linha_atual}"] = "Score de Compatibilidade:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            celula_score = ws[f"B{linha_atual}"]
            celula_score.value = f"{score}/100"
            if score >= 70:
                celula_score.font = Font(name="Arial", size=10, color="006100", bold=True)
            elif score >= 50:
                celula_score.font = Font(name="Arial", size=10, color="FFC000", bold=True)
            else:
                celula_score.font = Font(name="Arial", size=10, color="C00000", bold=True)
            celula_score.border = border
            linha_atual += 1
        
        # Análise
        analise_texto = avaliacao_cnae.get("analise", "")
        if analise_texto:
            ws[f"A{linha_atual}"] = "Análise:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = analise_texto
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            linha_atual += 1
        
        # Observações
        observacoes = avaliacao_cnae.get("observacoes", [])
        if observacoes:
            ws[f"A{linha_atual}"] = "Observações:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            linha_atual += 1
            
            for i, obs in enumerate(observacoes, 1):
                ws[f"B{linha_atual}"] = f"{i}. {obs}"
                ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
                ws[f"B{linha_atual}"].font = estilo_normal
                ws[f"B{linha_atual}"].border = border
                ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
                linha_atual += 1
        
        linha_atual += 1
    
    # SEÇÃO 4: ANÁLISE DE EMAIL E DOMÍNIO
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec4_email = ws[f"A{linha_atual}"]
    celula_sec4_email.value = "4. ANÁLISE DE EMAIL E DOMÍNIO"
    celula_sec4_email.font = estilo_cabecalho
    celula_sec4_email.fill = fill_cabecalho
    celula_sec4_email.alignment = alinhamento_esquerda
    linha_atual += 1
    
    # Email cadastrado vs CNPJA
    email_cadastrado = dados_empresa.get("email_cadastrado")
    email_cnpja = dados_empresa.get("email_cnpja")
    
    ws[f"A{linha_atual}"] = "Email Cadastrado:"
    ws[f"A{linha_atual}"].font = estilo_cabecalho
    ws[f"A{linha_atual}"].fill = fill_cinza
    ws[f"A{linha_atual}"].border = border
    
    ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
    ws[f"B{linha_atual}"] = email_cadastrado if email_cadastrado else "Não informado"
    ws[f"B{linha_atual}"].font = estilo_normal
    ws[f"B{linha_atual}"].border = border
    linha_atual += 1
    
    ws[f"A{linha_atual}"] = "Email CNPJA:"
    ws[f"A{linha_atual}"].font = estilo_cabecalho
    ws[f"A{linha_atual}"].fill = fill_cinza
    ws[f"A{linha_atual}"].border = border
    
    ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
    ws[f"B{linha_atual}"] = email_cnpja if email_cnpja else "Não encontrado"
    ws[f"B{linha_atual}"].font = estilo_normal
    ws[f"B{linha_atual}"].border = border
    linha_atual += 1
    
    # Comparação de domínios
    dominio_cadastro = dados_empresa.get("dominio_cadastro")
    dominio_cnpja = dados_empresa.get("dominio_cnpja")
    
    if dominio_cadastro and dominio_cnpja:
        dominios_iguais = dominio_cadastro.lower() == dominio_cnpja.lower()
        ws[f"A{linha_atual}"] = "Domínios Compatíveis:"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        celula_dominios = ws[f"B{linha_atual}"]
        if dominios_iguais:
            celula_dominios.value = f"✅ Sim ({dominio_cadastro})"
            celula_dominios.font = Font(name="Arial", size=10, color="006100")
        else:
            celula_dominios.value = f"❌ Não - Cadastro: {dominio_cadastro} | CNPJA: {dominio_cnpja}"
            celula_dominios.font = Font(name="Arial", size=10, color="C00000", bold=True)
        celula_dominios.border = border
        linha_atual += 1
    
    # Flags de risco de email
    sinalizacoes_email = []
    
    if dados_empresa.get("email_dominio_diferente"):
        sinalizacoes_email.append("❌ Email com domínio diferente do CNPJA")
    
    if dados_empresa.get("email_nao_corporativo"):
        sinalizacoes_email.append("⚠️ Email com domínio não corporativo (Gmail, Yahoo, etc.)")
    
    if dados_empresa.get("email_dominio_recente"):
        sinalizacoes_email.append("⚠️ Domínio do email criado recentemente (WHOIS)")
    
    if dados_empresa.get("email_typosquatting"):
        sinalizacoes_email.append("🚨 Possível typosquatting detectado (domínio similar ao CNPJA)")
    
    if sinalizacoes_email:
        ws[f"A{linha_atual}"] = "Sinalizações de Risco (Email):"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
        
        for sinalizacao in sinalizacoes_email:
            ws[f"B{linha_atual}"] = sinalizacao
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"].font = Font(name="Arial", size=10, color="C00000")
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
    else:
        ws[f"A{linha_atual}"] = "Sinalizações de Risco (Email):"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        ws[f"B{linha_atual}"] = "✅ Nenhuma sinalização de risco detectada"
        ws[f"B{linha_atual}"].font = Font(name="Arial", size=10, color="006100")
        ws[f"B{linha_atual}"].border = border
        linha_atual += 1
    
    # Detalhes da Idade do Domínio (WHOIS)
    whois_info = dados_empresa.get("whois_info")
    if whois_info and not whois_info.get("error"):
        linha_atual += 1
        ws[f"A{linha_atual}"] = "Detalhes da Idade do Domínio (WHOIS):"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
        
        # Data de criação
        creation_date = whois_info.get("creation_date")
        if creation_date:
            if isinstance(creation_date, str):
                try:
                    creation_date = datetime.fromisoformat(creation_date.replace('Z', '+00:00'))
                except:
                    pass
            if isinstance(creation_date, datetime):
                ws[f"A{linha_atual}"] = "Data de Criação:"
                ws[f"A{linha_atual}"].font = estilo_cabecalho
                ws[f"A{linha_atual}"].fill = fill_cinza
                ws[f"A{linha_atual}"].border = border
                
                ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
                ws[f"B{linha_atual}"] = creation_date.strftime("%d/%m/%Y")
                ws[f"B{linha_atual}"].font = estilo_normal
                ws[f"B{linha_atual}"].border = border
                linha_atual += 1
        
        # Idade em dias
        age_days = whois_info.get("age_days")
        if age_days is not None:
            ws[f"A{linha_atual}"] = "Idade do Domínio:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            celula_idade = ws[f"B{linha_atual}"]
            celula_idade.value = f"{age_days} dias"
            if age_days < 180:
                celula_idade.font = Font(name="Arial", size=10, color="C00000", bold=True)
            else:
                celula_idade.font = estilo_normal
            celula_idade.border = border
            linha_atual += 1
        
        # Limite configurado
        threshold_days = whois_info.get("threshold_days")
        if threshold_days is not None:
            ws[f"A{linha_atual}"] = "Limite Configurado:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = f"{threshold_days} dias"
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            linha_atual += 1
    
    # Detecção de Typosquatting (detalhes)
    typosquatting_info = dados_empresa.get("typosquatting_info")
    if typosquatting_info and typosquatting_info.get("suspeito"):
        linha_atual += 1
        ws[f"A{linha_atual}"] = "Detecção de Typosquatting:"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
        
        # Similaridade
        similaridade = typosquatting_info.get("similaridade", 0)
        ws[f"A{linha_atual}"] = "Similaridade:"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        celula_sim = ws[f"B{linha_atual}"]
        celula_sim.value = f"{similaridade:.1%}"
        celula_sim.font = Font(name="Arial", size=10, color="C00000", bold=True)
        celula_sim.border = border
        linha_atual += 1
        
        # Distância de Levenshtein
        distancia = typosquatting_info.get("distancia_levenshtein")
        if distancia is not None:
            ws[f"A{linha_atual}"] = "Distância (Levenshtein):"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = f"{distancia} caracteres"
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            linha_atual += 1
        
        # Typos detectados
        typos = typosquatting_info.get("typos_detectados", [])
        if typos:
            ws[f"A{linha_atual}"] = "Typos Detectados:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            linha_atual += 1
            
            for typo in typos:
                ws[f"B{linha_atual}"] = f"• {typo}"
                ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
                ws[f"B{linha_atual}"].font = estilo_normal
                ws[f"B{linha_atual}"].border = border
                ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
                linha_atual += 1
        
        # Mensagem
        mensagem = typosquatting_info.get("mensagem", "")
        if mensagem:
            ws[f"A{linha_atual}"] = "Análise:"
            ws[f"A{linha_atual}"].font = estilo_cabecalho
            ws[f"A{linha_atual}"].fill = fill_cinza
            ws[f"A{linha_atual}"].border = border
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = mensagem
            ws[f"B{linha_atual}"].font = Font(name="Arial", size=10, color="C00000")
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 4.5: OUTRAS SINALIZAÇÕES DE RISCO
    outras_sinalizacoes = []
    if dados_empresa.get("telefone_suspeito"):
        outras_sinalizacoes.append("❌ Telefone suspeito")
    if dados_empresa.get("pressa_aprovacao"):
        outras_sinalizacoes.append("⚠️ Pressa em aprovação")
    if dados_empresa.get("entrega_marcada"):
        outras_sinalizacoes.append("⚠️ Entrega marcada")
    if dados_empresa.get("endereco_entrega_diferente"):
        outras_sinalizacoes.append("⚠️ Endereço de entrega diferente do cadastro")
    
    if outras_sinalizacoes:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_sec45 = ws[f"A{linha_atual}"]
        celula_sec45.value = "4.5. OUTRAS SINALIZAÇÕES DE RISCO"
        celula_sec45.font = estilo_cabecalho
        celula_sec45.fill = fill_cabecalho
        celula_sec45.alignment = alinhamento_esquerda
        linha_atual += 1
        
        for sinalizacao in outras_sinalizacoes:
            ws[f"A{linha_atual}"] = sinalizacao
            ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
            ws[f"A{linha_atual}"].font = Font(name="Arial", size=10, color="C00000")
            ws[f"A{linha_atual}"].border = border
            ws[f"A{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
    else:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_sec45 = ws[f"A{linha_atual}"]
        celula_sec45.value = "4.5. OUTRAS SINALIZAÇÕES DE RISCO"
        celula_sec45.font = estilo_cabecalho
        celula_sec45.fill = fill_cabecalho
        celula_sec45.alignment = alinhamento_esquerda
        linha_atual += 1
        
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        ws[f"A{linha_atual}"] = "✅ Nenhuma outra sinalização de risco detectada"
        ws[f"A{linha_atual}"].font = Font(name="Arial", size=10, color="006100")
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 5: RESULTADO DA ANÁLISE DE RISCO
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec5 = ws[f"A{linha_atual}"]
    celula_sec5.value = "5. RESULTADO DA ANÁLISE DE RISCO"
    celula_sec5.font = estilo_cabecalho
    celula_sec5.fill = fill_cabecalho
    celula_sec5.alignment = alinhamento_esquerda
    linha_atual += 1
    
    # Risco Final (destaque)
    risco_final = analise_risco.get("risco_final", "INDEFINIDO")
    score_risco = analise_risco.get("score_risco", 0)
    
    ws.merge_cells(f"A{linha_atual}:B{linha_atual}")
    ws[f"A{linha_atual}"] = "RISCO FINAL:"
    ws[f"A{linha_atual}"].font = estilo_cabecalho
    ws[f"A{linha_atual}"].fill = fill_cinza
    ws[f"A{linha_atual}"].border = border
    ws[f"A{linha_atual}"].alignment = alinhamento_esquerda
    
    ws.merge_cells(f"C{linha_atual}:D{linha_atual}")
    celula_risco = ws[f"C{linha_atual}"]
    celula_risco.value = f"{risco_final} (Score: {score_risco}/100)"
    
    if risco_final == "ALTO":
        celula_risco.font = estilo_risco_alto
        celula_risco.fill = fill_risco_alto
    elif risco_final == "MEDIO":
        celula_risco.font = estilo_risco_medio
        celula_risco.fill = fill_risco_medio
    elif risco_final == "BAIXO":
        celula_risco.font = estilo_risco_baixo
        celula_risco.fill = fill_risco_baixo
    else:
        celula_risco.font = estilo_normal
    
    celula_risco.border = border
    celula_risco.alignment = alinhamento_centro
    linha_atual += 1
    
    # Tipo Local Esperado
    tipo_local_esperado = analise_risco.get("tipo_local_esperado", "N/A")
    ws[f"A{linha_atual}"] = "Tipo Local Esperado (CNAE):"
    ws[f"A{linha_atual}"].font = estilo_cabecalho
    ws[f"A{linha_atual}"].fill = fill_cinza
    ws[f"A{linha_atual}"].border = border
    
    ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
    ws[f"B{linha_atual}"] = tipo_local_esperado
    ws[f"B{linha_atual}"].font = estilo_normal
    ws[f"B{linha_atual}"].border = border
    linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 6: ANÁLISE VISUAL
    ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
    celula_sec5 = ws[f"A{linha_atual}"]
    celula_sec5.value = "6. ANÁLISE VISUAL (GEMINI VISION)"
    celula_sec5.font = estilo_cabecalho
    celula_sec5.fill = fill_cabecalho
    celula_sec5.alignment = alinhamento_esquerda
    linha_atual += 1
    
    analise_visual = analise_risco.get("analise_visual", {})
    
    dados_analise_visual = [
        ("Zona Aparente", analise_visual.get("zona_aparente", "N/A")),
        ("Tipo de Via", analise_visual.get("tipo_via", "N/A")),
        ("Placas Comerciais", "Sim" if analise_visual.get("presenca_placas_comerciais") else "Não"),
        ("Vitrines/Lojas", "Sim" if analise_visual.get("presenca_vitrines_ou_lojas") else "Não"),
        ("Casas Residenciais", "Sim" if analise_visual.get("presenca_casas_residenciais") else "Não"),
        ("Compatibilidade CNAE", analise_visual.get("compatibilidade_cnae", "N/A")),
        ("Sugestão de Risco", analise_visual.get("sugestao_nivel_risco", "N/A")),
    ]
    
    for rotulo, valor in dados_analise_visual:
        ws[f"A{linha_atual}"] = rotulo
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        ws[f"A{linha_atual}"].alignment = alinhamento_esquerda
        
        ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
        ws[f"B{linha_atual}"] = valor
        ws[f"B{linha_atual}"].font = estilo_normal
        ws[f"B{linha_atual}"].border = border
        ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
        linha_atual += 1
    
    # Motivos de Incompatibilidade
    motivos = analise_visual.get("motivos_incompatibilidade", [])
    if motivos:
        linha_atual += 1
        ws[f"A{linha_atual}"] = "Motivos de Incompatibilidade:"
        ws[f"A{linha_atual}"].font = estilo_cabecalho
        ws[f"A{linha_atual}"].fill = fill_cinza
        ws[f"A{linha_atual}"].border = border
        linha_atual += 1
        
        for i, motivo in enumerate(motivos, 1):
            ws[f"B{linha_atual}"] = f"{i}. {motivo}"
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
    
    linha_atual += 1
    
    # SEÇÃO 7: FLAGS DE RISCO
    flags = analise_risco.get("flags_risco", [])
    if flags:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_sec6 = ws[f"A{linha_atual}"]
        celula_sec6.value = "7. FLAGS DE RISCO"
        celula_sec6.font = estilo_cabecalho
        celula_sec6.fill = fill_cabecalho
        celula_sec6.alignment = alinhamento_esquerda
        linha_atual += 1
        
        for i, flag in enumerate(flags, 1):
            ws[f"A{linha_atual}"] = f"{i}."
            ws[f"A{linha_atual}"].font = estilo_normal
            ws[f"A{linha_atual}"].border = border
            ws[f"A{linha_atual}"].alignment = alinhamento_centro
            
            ws.merge_cells(f"B{linha_atual}:D{linha_atual}")
            ws[f"B{linha_atual}"] = flag
            ws[f"B{linha_atual}"].font = estilo_normal
            ws[f"B{linha_atual}"].border = border
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            linha_atual += 1
        
        linha_atual += 1
    
    # SEÇÃO 8: ANÁLISE DETALHADA
    analise_detalhada = analise_visual.get("analise_detalhada", "")
    if analise_detalhada:
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_sec7 = ws[f"A{linha_atual}"]
        celula_sec7.value = "8. ANÁLISE DETALHADA"
        celula_sec7.font = estilo_cabecalho
        celula_sec7.fill = fill_cabecalho
        celula_sec7.alignment = alinhamento_esquerda
        linha_atual += 1
        
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_detalhada = ws[f"A{linha_atual}"]
        celula_detalhada.value = analise_detalhada
        celula_detalhada.font = estilo_normal
        celula_detalhada.border = border
        celula_detalhada.alignment = alinhamento_esquerda
        celula_detalhada.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        linha_atual += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    
    # Salvar ou retornar bytes
    if caminho_saida:
        wb.save(caminho_saida)
        return None
    else:
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def gerar_relatorio_para_cnpj(cnpj: str, caminho_saida: Optional[str] = None) -> Optional[bytes]:
    """
    Gera relatório Excel completo para um CNPJ, buscando todos os dados necessários.
    
    Args:
        cnpj: CNPJ da empresa
        caminho_saida: Caminho para salvar o arquivo (opcional)
    
    Returns:
        Bytes do arquivo Excel ou None se caminho_saida fornecido
    """
    from database import (
        get_consulta_cnpj, get_endereco_geocoding, get_analise_risco_endereco,
        get_email_cnpja, get_dominio_email, get_avaliacao_cnae, get_config_whois_min_days
    )
    from whois_check import check_domain_age
    from typosquatting_detector import detect_typosquatting
    import sqlite3
    
    # Buscar dados
    dados_cnpj = get_consulta_cnpj(cnpj)
    dados_endereco = get_endereco_geocoding(cnpj)
    analise_risco = get_analise_risco_endereco(cnpj)
    avaliacao_cnae = get_avaliacao_cnae(cnpj)
    
    if not dados_cnpj:
        raise ValueError("CNPJ não encontrado no banco de dados")
    
    if not analise_risco:
        raise ValueError("Análise de risco não encontrada. Execute a análise primeiro.")
    
    # Buscar dados da empresa cadastrada (flags de email e outras sinalizações)
    # O CNPJ pode estar salvo com ou sem formatação, então vamos tentar ambas as formas
    cnpj_clean = "".join(filter(str.isdigit, cnpj))
    # Formatar CNPJ para buscar também com formatação
    cnpj_formatted = f"{cnpj_clean[:2]}.{cnpj_clean[2:5]}.{cnpj_clean[5:8]}/{cnpj_clean[8:12]}-{cnpj_clean[12:]}" if len(cnpj_clean) == 14 else cnpj
    
    conn = sqlite3.connect("savic.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Buscar todas as flags e sinalizações
    cursor.execute("""
        SELECT email, email_dominio_diferente, email_nao_corporativo, email_dominio_recente,
               email_typosquatting, telefone_suspeito, pressa_aprovacao, entrega_marcada, endereco_entrega_diferente
        FROM empresas
        WHERE cnpj = ? OR cnpj = ?
    """, (cnpj_formatted, cnpj_clean))
    
    empresa_row = cursor.fetchone()
    conn.close()
    
    # Preparar dados da empresa
    email_cadastrado = None
    email_dominio_diferente = False
    email_nao_corporativo = False
    email_dominio_recente = False
    email_typosquatting = False
    telefone_suspeito = False
    pressa_aprovacao = False
    entrega_marcada = False
    endereco_entrega_diferente = False
    
    if empresa_row:
        email_cadastrado = empresa_row["email"] if empresa_row["email"] and empresa_row["email"].strip() else None
        email_dominio_diferente = bool(empresa_row["email_dominio_diferente"]) if empresa_row["email_dominio_diferente"] is not None else False
        email_nao_corporativo = bool(empresa_row["email_nao_corporativo"]) if empresa_row["email_nao_corporativo"] is not None else False
        email_dominio_recente = bool(empresa_row["email_dominio_recente"]) if empresa_row["email_dominio_recente"] is not None else False
        email_typosquatting = bool(empresa_row["email_typosquatting"]) if empresa_row["email_typosquatting"] is not None else False
        telefone_suspeito = bool(empresa_row["telefone_suspeito"]) if empresa_row["telefone_suspeito"] is not None else False
        pressa_aprovacao = bool(empresa_row["pressa_aprovacao"]) if empresa_row["pressa_aprovacao"] is not None else False
        entrega_marcada = bool(empresa_row["entrega_marcada"]) if empresa_row["entrega_marcada"] is not None else False
        endereco_entrega_diferente = bool(empresa_row["endereco_entrega_diferente"]) if empresa_row["endereco_entrega_diferente"] is not None else False
    
    # Buscar email do CNPJA para comparação
    email_cnpja = get_email_cnpja(cnpj_clean)
    dominio_cadastro = get_dominio_email(email_cadastrado) if email_cadastrado else None
    dominio_cnpja = get_dominio_email(email_cnpja) if email_cnpja else None
    
    # Detectar typosquatting (mesmo que já tenha flag, vamos recalcular para ter detalhes)
    typosquatting_info = None
    if dominio_cadastro and dominio_cnpja and dominio_cadastro != dominio_cnpja:
        typosquatting_info = detect_typosquatting(dominio_cadastro, dominio_cnpja)
    
    # Buscar dados WHOIS detalhados
    whois_info = None
    if email_cadastrado:
        try:
            whois_info = check_domain_age(email_cadastrado)
        except Exception as e:
            print(f"Erro ao buscar dados WHOIS: {e}")
    
    dados_empresa = {
        "razao_social": dados_cnpj.get("company", {}).get("name") or dados_cnpj.get("name", "N/A"),
        "nome_fantasia": dados_cnpj.get("alias", "N/A"),
        "data_abertura": dados_cnpj.get("founded", "N/A"),
        "email_cadastrado": email_cadastrado,
        "email_cnpja": email_cnpja,
        "dominio_cadastro": dominio_cadastro,
        "dominio_cnpja": dominio_cnpja,
        "email_dominio_diferente": email_dominio_diferente,
        "email_nao_corporativo": email_nao_corporativo,
        "email_dominio_recente": email_dominio_recente,
        "email_typosquatting": email_typosquatting,
        "typosquatting_info": typosquatting_info,
        "whois_info": whois_info,
        "telefone_suspeito": telefone_suspeito,
        "pressa_aprovacao": pressa_aprovacao,
        "entrega_marcada": entrega_marcada,
        "endereco_entrega_diferente": endereco_entrega_diferente,
        "avaliacao_cnae": avaliacao_cnae
    }
    
    # Preparar CNAEs
    cnaes = []
    if dados_cnpj.get("mainActivity"):
        cnae_principal = dados_cnpj["mainActivity"]
        cnae_id = str(cnae_principal.get("id", ""))
        cnae_codigo = cnae_id if len(cnae_id) <= 7 else cnae_id[:7]
        cnaes.append({
            "codigo": cnae_codigo,
            "descricao": cnae_principal.get("text", "")
        })
    
    if dados_cnpj.get("sideActivities"):
        for sec in dados_cnpj["sideActivities"]:
            cnae_id = str(sec.get("id", ""))
            cnae_codigo = cnae_id if len(cnae_id) <= 7 else cnae_id[:7]
            cnaes.append({
                "codigo": cnae_codigo,
                "descricao": sec.get("text", "")
            })
    
    # Gerar relatório
    return gerar_relatorio_excel(
        cnpj=cnpj,
        dados_empresa=dados_empresa,
        dados_endereco=dados_endereco,
        analise_risco=analise_risco,
        cnaes=cnaes,
        caminho_saida=caminho_saida
    )

